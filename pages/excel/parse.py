import openpyxl as xl

STOCK = "在庫"
STOCK_KEY = "在庫数"

BOOK_NAME = "pages/excel/application.xlsx"

wb = xl.load_workbook(filename=BOOK_NAME)

stock_sheet = wb[STOCK]


def get_sheet_values(sheet):
    last = {"row": sheet.max_row, "col": sheet.max_column}
    sheet_values = []

    for row in range(last["row"]):
        sheet_values.append([])
        for col in range(last["col"]):
            value = sheet.cell(row + 1, col + 1).value
            if value is not None:
                sheet_values[row].append(value)
            pass
        pass

    return sheet_values


stock_values = get_sheet_values(stock_sheet)

labels = [label for label in stock_values[0]]

sheet_data = [
    {f"{label}": stock_values[value_i][label_i] for label_i, label in enumerate(labels)}
    for value_i in range(1, len(stock_values))
]
for datum in sheet_data:
    print(datum)


def set_sheet_values(data: list[dict[str, str]]):
    keys = list(data[0].keys())
    stock_i = keys.index(STOCK_KEY)

    prev_stocks = []

    for row, datum in enumerate(data):
        cell = stock_sheet.cell(row=row + 2, column=stock_i + 1)

        prev_stocks.append(cell.value)
        cell.value = datum[STOCK_KEY]
        pass

    # wb.save(BOOK_NAME)

    return sum(
        [
            int(datum["単価"]) * (int(prev_stocks[i]) - int(datum[STOCK_KEY]))
            for i, datum in enumerate(data)
        ]
    )
